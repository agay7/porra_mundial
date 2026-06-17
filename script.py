import pandas as pd
from pathlib import Path

# ======================
# CONFIGURACIÓN
# ======================
RUTA_MAESTRO = Path("data/maestro.xlsx")
RUTA_PARTICIPANTES = Path("data/participantes")
SALIDA = "clasificacion.xlsx"

# PUNTOS
P_GANADOR = 1
P_DIFERENCIA = 3
P_RESULTADO_EXACTO = 5
P_POSICION_GRUPO = 3
P_PICHICHI = 10
P_BALON_ORO = 10


# ======================
# PARTIDOS
# ======================
def ganador(gl, gv):
    if gl > gv:
        return 1
    elif gl < gv:
        return -1
    return 0


def acierto_equipos(real, apuesta):
    local_r = str(real["LOCAL"]).strip().upper()
    visitante_r = str(real["VISITANTE"]).strip().upper()

    local_a = str(apuesta["LOCAL"]).strip().upper()
    visitante_a = str(apuesta["VISITANTE"]).strip().upper()

    equipos_real = {local_r, visitante_r}
    equipos_apuesta = {local_a, visitante_a}

    return len(equipos_real.intersection(equipos_apuesta))  # 0, 1 o 2


def puntos_partido(real, apuesta):
    gl_r = real["GOLES LOCAL"]
    gv_r = real["GOLES VISITANTE"]
    gl_a = apuesta["GOLES LOCAL"]
    gv_a = apuesta["GOLES VISITANTE"]

    # ✅ nombres equipos reales y apuesta
    local_r = str(real["LOCAL"]).strip().upper()
    visitante_r = str(real["VISITANTE"]).strip().upper()
    local_a = str(apuesta["LOCAL"]).strip().upper()
    visitante_a = str(apuesta["VISITANTE"]).strip().upper()

    # ✅ comprobar equipos acertados
    aciertos = acierto_equipos(real, apuesta)

    # ❌ no acierta ningún equipo → 0
    if aciertos == 0:
        return 0

    # ✅ determinar equipo ganador REAL
    if gl_r > gv_r:
        ganador_real_equipo = local_r
    elif gv_r > gl_r:
        ganador_real_equipo = visitante_r
    else:
        ganador_real_equipo = None  # empate

    # ✅ determinar equipo ganador APUESTA
    if gl_a > gv_a:
        ganador_apuesta_equipo = local_a
    elif gv_a > gl_a:
        ganador_apuesta_equipo = visitante_a
    else:
        ganador_apuesta_equipo = None

    # ❌ si no acierta el equipo ganador REAL → 0
    if ganador_real_equipo != ganador_apuesta_equipo:
        return 0

    # ✅ resultado exacto
    if gl_r == gl_a and gv_r == gv_a:
        return P_RESULTADO_EXACTO

    # ✅ diferencia (solo si acierta ganador)
    if (gl_r - gv_r) == (gl_a - gv_a):
        return P_DIFERENCIA

    # ✅ solo ganador (incluye clasificado)
    return P_GANADOR


def puntuar_partidos(maestro, jugador):
    total = 0

    ganadores = 0
    diferencias = 0
    exactos = 0

    partidos = maestro[(maestro["ID"] >= 1) & (maestro["ID"] <= 104)]

    for _, real in partidos.iterrows():

        if int(real.get("JUGADO", 0)) != 1:
            continue

        fila = jugador[jugador["ID"] == real["ID"]]
        if fila.empty:
            continue

        apuesta = fila.iloc[0]

        puntos = puntos_partido(real, apuesta)

        total += puntos

        if puntos == P_RESULTADO_EXACTO:
            exactos += 1

        elif puntos == P_DIFERENCIA:
            diferencias += 1

        elif puntos == P_GANADOR:
            ganadores += 1

    return total, ganadores, diferencias, exactos


# ======================
# GRUPOS
# ======================
def normalizar_grupo(valor):
    v = str(valor).strip().upper()
    if v.startswith("GRUPO") and "_" not in v:
        return "GRUPO_" + v.replace("GRUPO", "")
    return v


def leer_grupos_desde_datos(df):
    grupos = {}

    filas = df[
        (df["ID"] >= 1000)
        & (df["LOCAL"].astype(str).str.upper().str.startswith("GRUPO"))
    ]

    for _, fila in filas.iterrows():
        grupo = normalizar_grupo(fila["LOCAL"])
        posicion = str(fila["VISITANTE"]).strip()
        clave = f"{grupo}_{posicion}"
        grupos[clave] = str(fila["GOLES LOCAL"]).strip()

    return grupos


def puntuar_grupos(maestro, jugador):
    puntos = 0
    for k, v in maestro.items():
        if jugador.get(k) == v:
            puntos += P_POSICION_GRUPO
    return puntos


def grupos_terminados(maestro):
    fila = maestro[maestro["ID"] == 72]
    if fila.empty:
        return False
    return int(fila.iloc[0].get("JUGADO", 0)) == 1


# ======================
# PREMIOS
# ======================
def leer_premios_maestro(df):
    premios = {}

    fila_pichichi = df[df["ID"] == 988]
    if not fila_pichichi.empty:
        premios["Pichichi"] = {
            "valor": str(fila_pichichi.iloc[0]["VISITANTE"]).strip(),
            "jugado": int(fila_pichichi.iloc[0]["JUGADO"])
        }

    fila_balon = df[df["ID"] == 999]
    if not fila_balon.empty:
        premios["Balón de Oro"] = {
            "valor": str(fila_balon.iloc[0]["VISITANTE"]).strip(),
            "jugado": int(fila_balon.iloc[0]["JUGADO"])
        }

    return premios


def leer_premios_jugador(df):
    premios = {}

    fila_pichichi = df[df["ID"] == 988]
    if not fila_pichichi.empty:
        premios["Pichichi"] = str(fila_pichichi.iloc[0]["VISITANTE"]).strip()

    fila_balon = df[df["ID"] == 999]
    if not fila_balon.empty:
        premios["Balón de Oro"] = str(fila_balon.iloc[0]["VISITANTE"]).strip()

    return premios


def puntuar_premios(maestro, jugador):
    puntos = 0

    if (
        maestro["Pichichi"]["jugado"] == 1
        and jugador.get("Pichichi")
        and jugador["Pichichi"].lower()
        == maestro["Pichichi"]["valor"].lower()
    ):
        puntos += P_PICHICHI

    if (
        maestro["Balón de Oro"]["jugado"] == 1
        and jugador.get("Balón de Oro")
        and jugador["Balón de Oro"].lower()
        == maestro["Balón de Oro"]["valor"].lower()
    ):
        puntos += P_BALON_ORO

    return puntos


# ======================
# PROGRESO GLOBAL
# ======================
def contar_partidos_jugados(maestro):
    partidos = maestro[(maestro["ID"] >= 1) & (maestro["ID"] <= 104)]
    return int((partidos["JUGADO"] == 1).sum())


# ======================
# MAIN
# ======================
def main():
    print("📊 Calculando clasificación de la porra...\n")

    maestro_datos = pd.read_excel(RUTA_MAESTRO, sheet_name="Datos", engine="openpyxl")

    partidos_jugados = contar_partidos_jugados(maestro_datos)
    partidos_totales = 104

    print("📊 PROGRESO DEL TORNEO")
    print(f"Partidos jugados: {partidos_jugados} / {partidos_totales}\n")

    maestro_grupos = leer_grupos_desde_datos(maestro_datos)
    maestro_premios = leer_premios_maestro(maestro_datos)

    import os

    clasificacion_anterior = None

    if os.path.exists(SALIDA):
        try:
            clasificacion_anterior = pd.read_excel(SALIDA)
        except:
            pass

    ranking = []

    for archivo in RUTA_PARTICIPANTES.glob("*.xlsx"):
        if archivo.name.startswith("~$"):
            continue

        nombre = archivo.stem
        print(f"➡ Procesando: {nombre}")

        jugador_datos = pd.read_excel(archivo, sheet_name="Datos", engine="openpyxl")
        jugador_grupos = leer_grupos_desde_datos(jugador_datos)
        jugador_premios = leer_premios_jugador(jugador_datos)

        puntos_partidos, ganadores, diferencias, exactos = puntuar_partidos(
            maestro_datos,
            jugador_datos
        )

        if grupos_terminados(maestro_datos):
            puntos_grupos = puntuar_grupos(maestro_grupos, jugador_grupos)
        else:
            puntos_grupos = 0

        puntos_premios = puntuar_premios(maestro_premios, jugador_premios)

        ranking.append({
            "Participante": nombre,
            "Puntos por partidos": puntos_partidos,
            "Signo (1X2)": ganadores,
            "Diferencia de goles": diferencias,
            "Resultados exactos": exactos,
            "Puntos por grupos": puntos_grupos,
            "Puntos por premios": puntos_premios,
            "Totales": puntos_partidos + puntos_grupos + puntos_premios
        })

    df = (
        pd.DataFrame(ranking)
        .sort_values("Totales", ascending=False)
        .reset_index(drop=True)
    )

    df.insert(0, "Posición", df.index + 1)

    import os

    if os.path.exists(SALIDA):

        try:
            clasificacion_anterior = pd.read_excel(SALIDA)

            posiciones_antiguas = {
                fila["Jugador"]: fila["Posición"]
                for _, fila in clasificacion_anterior.iterrows()
            }

            movimientos = []

            for _, fila in df.iterrows():

                jugador = fila["Jugador"]
                posicion_actual = fila["Posición"]

                if jugador not in posiciones_antiguas:
                    movimientos.append("🆕")
                    continue

                posicion_anterior = posiciones_antiguas[jugador]

                diferencia = posicion_anterior - posicion_actual

                if diferencia > 0:
                    movimientos.append(f"↑{diferencia}")
                elif diferencia < 0:
                    movimientos.append(f"↓{abs(diferencia)}")
                else:
                    movimientos.append("=")

            df.insert(1, "Mov", movimientos)

        except Exception as e:
            print(f"No se pudo leer la clasificación anterior: {e}")

    if clasificacion_anterior is not None:

        posiciones_antiguas = {}

        for _, fila in clasificacion_anterior.iterrows():
            posiciones_antiguas[fila["Participante"]] = fila["Posición"]

        movimientos = []

        for _, fila in df.iterrows():

            jugador = fila["Participante"]
            posicion_actual = fila["Posición"]

            if jugador not in posiciones_antiguas:
                movimientos.append("🆕")
                continue

            posicion_anterior = posiciones_antiguas[jugador]

            diferencia = posicion_anterior - posicion_actual

            if diferencia > 0:
                movimientos.append(f"↑{diferencia}")

            elif diferencia < 0:
                movimientos.append(f"↓{abs(diferencia)}")

            else:
                movimientos.append("=")

        df.insert(1, "Movimiento", movimientos)

    print("🏆 CLASIFICACIÓN FINAL\n")
    print(df.to_string(index=False))

    df.to_excel(SALIDA, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(SALIDA)
    ws = wb.active

    verde = PatternFill("solid", fgColor="C6EFCE")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    gris = PatternFill("solid", fgColor="D9D9D9")

    # Buscar la columna Mov
    col_mov = None

    for cell in ws[1]:
        if cell.value == "Mov":
            col_mov = cell.column
            break

    if col_mov:

        for fila in range(2, ws.max_row + 1):

            celda = ws.cell(row=fila, column=col_mov)
            valor = str(celda.value)

            if valor.startswith("↑"):
                celda.fill = verde

            elif valor.startswith("↓"):
                celda.fill = rojo

            elif valor == "=":
                celda.fill = gris

    wb.save(SALIDA)

    print(f"\n✅ Clasificación guardada en {SALIDA}")
if __name__ == "__main__":
    main()